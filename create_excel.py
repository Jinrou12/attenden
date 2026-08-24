import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import csv

students = [
    "ឡុញ ចន្ទបុត្រ",
    "វិន ឈាងលី",
    "យិន លីហ្សា",
    "គាត តុលា",
    "ឌិត ឃុនឡេង",
    "រិត យូសេង",
    "គួន សុកគីន",
    "នឿន វ៉ាន់ដេត",
    "សាត សុធារិទ្ធ",
    "អេង ប៊ុនលី",
    "សល់ រស្មី",
    "សុខ ចាន់ណា",
    "ធី ពិសិដ្ឋ",
    "ម៉េន ពន្លក",
    "អឿន ឆារ៉ាន់់",
    "រឿន ម៉េងលាង",
    "ផល្លី ចាន់ឌី",
    "វន់ ដាវិត",
    "ស៊ុយ មេងលាង",
    "កាន គឹមហ័ង",
    "អ៊ុម កៅអង់",
    "អ៊ុត ខាត់នី",
    "ភិន ម៊ីុ",
    "រីម សុផានិត",
    "ហៀង ដាលីន",
    "សន សីលា",
    "នឿន វ៉ាន់រ៉េត",
    "ហ៊ន ហាប់",
    "ទូច ពិសី",
    "ផាត ម៉េងហ៊ាង",
    "សាំង ស៊ាន់",
    "ថែ អាល់",
    "ភឿន ប៊ុនថាវ",
    "កយ ហួត",
    "ថន ប៊ុនធូ",
    "អ៊ុល ឆៃអែង",
    "ភាព តារ៉ា",
    "ចឺម រតនា",
    "តូ ឡុងហេង",
    "ធឿន រក្សា",
    "ភាក់ សុផល",
    "ចាន់រ៉ា សុចិត្រា",
    "ថុន វិសាល"
]

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "បញ្ជីឈ្មោះសិស្ស ថ្នាក់ទី១២"

# Title Header
ws.merge_cells('A1:D1')
ws['A1'] = "បញ្ជីឈ្មោះសិស្ស ថ្នាក់ទី១២ (អនុវិទ្យាល័យ សម្តេច ជុន ណាត)"
ws['A1'].font = Font(name='Khmer OS Battambang', size=14, bold=True, color='1F2937')
ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws.row_dimensions[1].height = 35

# Column Headers
headers = ["ល.រ (No.)", "អត្តលេខ (Code)", "ឈ្មោះសិស្ស (Student Name)", "ភេទ (Gender)"]
ws.append([]) # Blank row 2

for col_num, header in enumerate(headers, 1):
    cell = ws.cell(row=3, column=col_num)
    cell.value = header
    cell.font = Font(name='Khmer OS Battambang', size=11, bold=True, color='FFFFFF')
    cell.fill = PatternFill(start_color='059669', end_color='059669', fill_type='solid')
    cell.alignment = Alignment(horizontal='center', vertical='center')

ws.row_dimensions[3].height = 25

thin_border = Border(
    left=Side(style='thin', color='CBD5E1'),
    right=Side(style='thin', color='CBD5E1'),
    top=Side(style='thin', color='CBD5E1'),
    bottom=Side(style='thin', color='CBD5E1')
)

csv_data = [["No", "Code", "Name", "Gender"]]

for idx, name in enumerate(students, 1):
    code = f"STU-{idx:03d}"
    gender = "ប្រុស"
    row_num = idx + 3
    
    ws.cell(row=row_num, column=1, value=idx).alignment = Alignment(horizontal='center')
    ws.cell(row=row_num, column=2, value=code).alignment = Alignment(horizontal='center')
    ws.cell(row=row_num, column=3, value=name).alignment = Alignment(horizontal='left')
    ws.cell(row=row_num, column=4, value=gender).alignment = Alignment(horizontal='center')
    
    for c in range(1, 5):
        cell = ws.cell(row=row_num, column=c)
        cell.font = Font(name='Khmer OS Battambang', size=10)
        cell.border = thin_border
        if idx % 2 == 0:
            cell.fill = PatternFill(start_color='F8FAFC', end_color='F8FAFC', fill_type='solid')
            
    csv_data.append([idx, code, name, gender])

ws.column_dimensions['A'].width = 12
ws.column_dimensions['B'].width = 16
ws.column_dimensions['C'].width = 28
ws.column_dimensions['D'].width = 14

wb.save("Student_List_Grade12.xlsx")

with open("Student_List_Grade12.csv", "w", encoding="utf-8-sig", newline="") as f:
    writer = csv.writer(f)
    writer.writerows(csv_data)

print("Updated Student_List_Grade12.xlsx and Student_List_Grade12.csv successfully!")
